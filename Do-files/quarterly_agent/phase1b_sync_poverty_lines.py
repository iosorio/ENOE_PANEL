#!/usr/bin/env python3
"""Phase 1B: sync INEGI poverty lines and patch target harmonization do-file.

Goals:
1) Build canonical monthly/quarterly poverty-line tables under Doc/poverty_lines_inegi/.
2) Use INEGI as the only upstream source.
3) Patch the target quarter harmonization do-file to dynamically set only the
   required scalar pair (urban/rural) for the current quarter.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import json
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

USER_AGENT = "ENOE-Quarterly-Agent/1.0 (+phase1b)"
INEGI_LP_PAGE = "https://www.inegi.org.mx/desarrollosocial/lp/"
INEGI_LP_JS = "https://www.inegi.org.mx/desarrollosocial/lp/js/contenido.min.js"

# INEGI indicator IDs used by LP component for extreme poverty line (canasta alimentaria).
INDICATOR_EXTREME_RURAL = "8999998849"
INDICATOR_EXTREME_URBAN = "8999998850"
INDICATOR_TOKEN = "96fbd1bf-21e6-28e3-6e64-2b15999d2c89"
INDICATOR_API_TPL = (
    "https://www.inegi.org.mx/app/api/indicadores/interna_v1_3/Indicador/"
    "{indicator}/00/es/false/null/json/{token}"
)

SPANISH_MONTHS = {
    "ene": 1,
    "feb": 2,
    "mar": 3,
    "abr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dic": 12,
}

POVERTY_BLOCK_PATTERN = re.compile(
    r"\*<_povertyincome_>.*?(?:\*</_povertyincome_>|</_povertyincome_>)",
    re.DOTALL,
)
POVERTY_BLOCK_REPLACEMENT = """*<_povertyincome_>
	gen ocupado=cond(clase1==1 & clase2==1,1,0)

	destring p6b2 p6c, replace
	recode p6b2 (999998=.) (999999=.)

	*Recuperacion de ingresos por rangos de salarios minimos
	gen double ingreso=p6b2
	replace ingreso=0 if ocupado==0
	replace ingreso=0 if p6b2==. & (p6_9==9 | p6a3==3)
	replace ingreso=0.5*salario if p6b2==. & p6c==1
	replace ingreso=1*salario if p6b2==. & p6c==2
	replace ingreso=1.5*salario if p6b2==. & p6c==3
	replace ingreso=2.5*salario if p6b2==. & p6c==4
	replace ingreso=4*salario if p6b2==. & p6c==5
	replace ingreso=7.5*salario if p6b2==. & p6c==6
	replace ingreso=10*salario if p6b2==. & p6c==7

	gen tamh = 1

	rename fac factor
	gen rururb = cond(t_loc>=1 & t_loc<=3,0,1)
	label define ru 0 "Urbano" 1 "Rural"
	label values rururb ru
	destring ent, replace

	gen mv=cond(ingreso==. & ocupado==1,1,0)

	foreach var in tamh ingreso mv ocupado {
		rename `var' _`var'
		bys folioh: egen double `var' = sum(_`var')
		drop _`var'
	}

	*Se elimina a los hogares que tienen valores perdidos en ingreso
	replace mv=1 if mv>0 & mv!=.
	*drop if mv==1

	gen _quarter = substr(wave,2,1)
	destring _quarter, replace
	sum _quarter
		local q = r(mean)
	drop _quarter
	sum year
		local yy = r(mean)
		local y = substr("`yy'",3,2)
	local x = `q'`y'
	noi di "local x = `x'"

	* Dynamic poverty lines (INEGI source, quarterly table)
	local poverty_csv "$path/Doc/poverty_lines_inegi/poverty_lines_quarterly.csv"
	cap confirm file "`poverty_csv'"
	if _rc {
		local poverty_csv "`server'/Doc/poverty_lines_inegi/poverty_lines_quarterly.csv"
		cap confirm file "`poverty_csv'"
	}
	if _rc {
		di as error "Missing poverty lines CSV: `poverty_csv'"
		exit 601
	}

	preserve
		import delimited using "`poverty_csv'", clear varnames(1)
		destring year quarter rural urban, replace force
		keep if year == `yy' & quarter == `q'
		count
		if r(N) != 1 {
			di as error "Missing poverty line row for `yy'-Q`q' in `poverty_csv'"
			exit 459
		}
		local lp_rural = rural[1]
		local lp_urban = urban[1]
	restore

	scalar uT`x' = `lp_urban'
	scalar rT`x' = `lp_rural'
	gen lpT`x' = cond(rururb==0,uT`x',rT`x')
	gen pob = cond((ingreso/tamh)<lpT`x',1,0)

	replace pob = . if mv==1
	drop rururb ocupado
	rename lpT`x' lpT

	label var tamh 		"Household size for per capita consumption, CONEVAL"
	label var ingreso 	"Income monthly LCU, adding p6c minimum wage brackets, CONEVAL"
	label var mv 		"Household with missing incomes excluded from poverty calculation, CONEVAL"
	label var lpT		"Income poverty line, CONEVAL"
	label var pob		"Poor by income, CONEVAL"
*</_povertyincome_>"""


@dataclass(frozen=True)
class MonthlyPoint:
    year: int
    month: int
    rural: float
    urban: float


@dataclass(frozen=True)
class QuarterlyPoint:
    year: int
    quarter: int
    rural: float
    urban: float
    month_count: int


def utc_now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


def timestamp_slug() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def parse_args() -> argparse.Namespace:
    script = Path(__file__).resolve()
    default_repo = script.parents[2]
    default_state_dir = default_repo / "Do-files" / "quarterly_agent" / "state" / "poverty"
    default_doc_dir = default_repo / "Doc" / "poverty_lines_inegi"

    ap = argparse.ArgumentParser(description="Sync INEGI poverty lines and patch target harmonization do-file")
    ap.add_argument("--repo-root", default=str(default_repo))
    ap.add_argument("--target-year", type=int, default=None)
    ap.add_argument("--target-quarter", type=int, choices=[1, 2, 3, 4], default=None)
    ap.add_argument("--source-mode", choices=["auto", "xlsx", "api"], default="auto")
    ap.add_argument("--timeout-seconds", type=int, default=45)
    ap.add_argument("--doc-dir", default=str(default_doc_dir))
    ap.add_argument("--state-dir", default=str(default_state_dir))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-do-patch", action="store_true")
    ap.add_argument("--verbose", action="store_true")
    return ap.parse_args()


def fetch_bytes(url: str, timeout_seconds: int) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=timeout_seconds) as resp:
        return resp.read()


def fetch_text(url: str, timeout_seconds: int) -> str:
    data = fetch_bytes(url, timeout_seconds=timeout_seconds)
    return data.decode("utf-8", errors="replace")


def request_json(url: str, timeout_seconds: int) -> Any:
    text = fetch_text(url, timeout_seconds=timeout_seconds)
    return json.loads(text)


def extract_urls_from_text(text: str, base_url: str) -> list[str]:
    out: set[str] = set()
    for m in re.finditer(r'https?://[^\s"\'\)]+(?:\.xlsx|\.xls|\.zip)', text, flags=re.IGNORECASE):
        out.add(m.group(0))
    for m in re.finditer(r'(/[^\s"\'\)]+(?:\.xlsx|\.xls|\.zip))', text, flags=re.IGNORECASE):
        out.add(urllib.parse.urljoin(base_url, m.group(1)))
    return sorted(out)


def _extract_date_score(url: str) -> tuple[int, int]:
    low = url.lower()

    m = re.search(r"(20\d{2})[_-](\d{2})", low)
    if m:
        return int(m.group(1)), int(m.group(2))

    m = re.search(r"(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)(20\d{2})", low)
    if m:
        return int(m.group(2)), SPANISH_MONTHS[m.group(1)]

    return 0, 0


def discover_inegi_xlsx_candidates(timeout_seconds: int, verbose: bool) -> list[str]:
    urls: set[str] = set()
    seeds = [INEGI_LP_PAGE, INEGI_LP_JS]

    for seed in seeds:
        try:
            text = fetch_text(seed, timeout_seconds=timeout_seconds)
        except Exception as exc:  # noqa: BLE001
            if verbose:
                print(f"WARN could not read {seed}: {exc}")
            continue
        urls.update(extract_urls_from_text(text, base_url=seed))

    # Keep INEGI-hosted files only (INEGI-only policy).
    inegi_urls = [u for u in urls if urllib.parse.urlparse(u).netloc.endswith("inegi.org.mx")]
    return sorted(inegi_urls, key=_extract_date_score, reverse=True)


def _find_best_sheet_name(wb: openpyxl.Workbook) -> str | None:
    for name in wb.sheetnames:
        low = name.lower()
        if "pobreza" in low and "ingresos" in low:
            return name
    return wb.sheetnames[0] if wb.sheetnames else None


def parse_monthly_from_xlsx_bytes(content: bytes) -> list[MonthlyPoint]:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ModuleNotFoundError as exc:  # pragma: no cover - env dependent
        raise RuntimeError(
            "openpyxl is not installed; cannot parse INEGI XLSX source in this environment"
        ) from exc

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet_name = _find_best_sheet_name(wb)
    if sheet_name is None:
        raise ValueError("xlsx has no worksheets")

    ws = wb[sheet_name]
    monthly: list[MonthlyPoint] = []

    for row in range(1, ws.max_row + 1):
        date_val = ws.cell(row=row, column=6).value
        rural = ws.cell(row=row, column=7).value
        urban = ws.cell(row=row, column=8).value
        if not isinstance(date_val, dt.datetime):
            continue
        if not isinstance(rural, (int, float)) or not isinstance(urban, (int, float)):
            continue
        monthly.append(
            MonthlyPoint(
                year=int(date_val.year),
                month=int(date_val.month),
                rural=float(rural),
                urban=float(urban),
            )
        )

    if not monthly:
        raise ValueError("xlsx parsed but no monthly records found")

    monthly.sort(key=lambda x: (x.year, x.month))
    return monthly


def parse_monthly_from_zip_bytes(content: bytes) -> tuple[list[MonthlyPoint], str]:
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xlsm"))]
        if not names:
            raise ValueError("zip does not contain xlsx files")
        names = sorted(names, key=lambda n: _extract_date_score(n), reverse=True)
        first = names[0]
        xlsx_bytes = zf.read(first)
    return parse_monthly_from_xlsx_bytes(xlsx_bytes), first


def parse_inegi_indicator_series(indicator_id: str, timeout_seconds: int) -> dict[tuple[int, int], float]:
    url = INDICATOR_API_TPL.format(indicator=indicator_id, token=INDICATOR_TOKEN)
    payload = request_json(url, timeout_seconds=timeout_seconds)

    if isinstance(payload, dict) and payload.get("ErrorCode"):
        raise RuntimeError(f"INEGI indicator API error for {indicator_id}: {payload.get('ErrorInfo')}")

    data = payload.get("Data", {}) if isinstance(payload, dict) else {}
    series = data.get("Serie", []) if isinstance(data, dict) else []
    if not isinstance(series, list) or not series:
        raise RuntimeError(f"No series data for indicator {indicator_id}")

    out: dict[tuple[int, int], float] = {}
    for row in series:
        if not isinstance(row, dict):
            continue
        period = str(row.get("TimePeriod", ""))
        value = str(row.get("CurrentValue", "")).replace(",", "")
        m = re.match(r"^(\d{4})/(\d{2})$", period)
        if not m:
            continue
        try:
            yy = int(m.group(1))
            mm = int(m.group(2))
            vv = float(value)
        except ValueError:
            continue
        out[(yy, mm)] = vv

    if not out:
        raise RuntimeError(f"Parsed empty series for indicator {indicator_id}")

    return out


def monthly_from_inegi_api(timeout_seconds: int) -> tuple[list[MonthlyPoint], dict[str, Any]]:
    rural = parse_inegi_indicator_series(INDICATOR_EXTREME_RURAL, timeout_seconds=timeout_seconds)
    urban = parse_inegi_indicator_series(INDICATOR_EXTREME_URBAN, timeout_seconds=timeout_seconds)

    keys = sorted(set(rural.keys()) & set(urban.keys()))
    monthly = [MonthlyPoint(year=yy, month=mm, rural=rural[(yy, mm)], urban=urban[(yy, mm)]) for (yy, mm) in keys]
    if not monthly:
        raise RuntimeError("No overlapping monthly points between rural and urban INEGI indicators")

    meta = {
        "type": "inegi_api",
        "indicator_ids": {
            "rural": INDICATOR_EXTREME_RURAL,
            "urban": INDICATOR_EXTREME_URBAN,
        },
        "api_template": INDICATOR_API_TPL,
    }
    return monthly, meta


def quantize_2(value: float) -> float:
    d = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(d)


def build_quarterly(monthly: list[MonthlyPoint]) -> list[QuarterlyPoint]:
    buckets: dict[tuple[int, int], list[MonthlyPoint]] = defaultdict(list)
    for row in monthly:
        q = ((row.month - 1) // 3) + 1
        buckets[(row.year, q)].append(row)

    out: list[QuarterlyPoint] = []
    for (yy, qq), rows in sorted(buckets.items()):
        rows = sorted(rows, key=lambda x: x.month)
        if len(rows) != 3:
            out.append(
                QuarterlyPoint(
                    year=yy,
                    quarter=qq,
                    rural=quantize_2(sum(r.rural for r in rows) / len(rows)),
                    urban=quantize_2(sum(r.urban for r in rows) / len(rows)),
                    month_count=len(rows),
                )
            )
            continue

        out.append(
            QuarterlyPoint(
                year=yy,
                quarter=qq,
                rural=quantize_2(sum(r.rural for r in rows) / 3.0),
                urban=quantize_2(sum(r.urban for r in rows) / 3.0),
                month_count=3,
            )
        )

    return out


def max_month_key(rows: list[MonthlyPoint]) -> tuple[int, int]:
    return max((row.year, row.month) for row in rows)


def write_monthly_csv(path: Path, rows: list[MonthlyPoint]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["year", "month", "rural", "urban"])
        for row in rows:
            w.writerow([row.year, row.month, f"{row.rural:.2f}", f"{row.urban:.2f}"])


def write_quarterly_csv(path: Path, rows: list[QuarterlyPoint]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["year", "quarter", "rural", "urban", "months"])
        for row in rows:
            if row.month_count != 3:
                continue
            w.writerow([row.year, row.quarter, f"{row.rural:.2f}", f"{row.urban:.2f}", row.month_count])


def read_monthly_csv(path: Path) -> list[MonthlyPoint]:
    if not path.exists():
        raise FileNotFoundError(f"Missing monthly CSV: {path}")

    out: list[MonthlyPoint] = []
    with path.open("r", encoding="utf-8", newline="") as fh:
        rdr = csv.DictReader(fh)
        required = {"year", "month", "rural", "urban"}
        if not rdr.fieldnames or not required.issubset(set(rdr.fieldnames)):
            raise ValueError(f"Monthly CSV missing required columns: {path}")
        for row in rdr:
            try:
                out.append(
                    MonthlyPoint(
                        year=int(row["year"]),
                        month=int(row["month"]),
                        rural=float(row["rural"]),
                        urban=float(row["urban"]),
                    )
                )
            except (TypeError, ValueError, KeyError):
                continue

    if not out:
        raise ValueError(f"Monthly CSV has no valid rows: {path}")
    out.sort(key=lambda x: (x.year, x.month))
    return out


def patch_target_harmonization_do(repo_root: Path, year: int, quarter: int, dry_run: bool) -> dict[str, Any]:
    qroot = repo_root / f"MEX_{year}_ENOE-Q{quarter}"
    do_path = (
        qroot
        / f"MEX_{year}_ENOE_V01_M_V06_A_GLD"
        / "Programs"
        / f"MEX_{year}_ENOE_V01_M_V06_A_GLD_ALL.do"
    )

    if not do_path.exists():
        return {"status": "missing_do", "path": str(do_path)}

    text = do_path.read_text(encoding="utf-8", errors="replace")
    replaced, n = POVERTY_BLOCK_PATTERN.subn(POVERTY_BLOCK_REPLACEMENT, text, count=1)
    if n != 1:
        return {
            "status": "patch_markers_not_found",
            "path": str(do_path),
            "message": "Could not locate povertyincome markers in harmonization do-file",
        }

    if replaced == text:
        return {"status": "already_patched", "path": str(do_path)}

    if not dry_run:
        do_path.write_text(replaced, encoding="utf-8")

    return {"status": "patched" if not dry_run else "would_patch", "path": str(do_path)}


def choose_best_monthly_from_xlsx(candidates: list[str], timeout_seconds: int, verbose: bool) -> tuple[list[MonthlyPoint], dict[str, Any]]:
    best_rows: list[MonthlyPoint] | None = None
    best_meta: dict[str, Any] | None = None
    best_key: tuple[int, int] = (0, 0)

    for url in candidates:
        try:
            raw = fetch_bytes(url, timeout_seconds=timeout_seconds)
            if url.lower().endswith(".zip"):
                rows, inner_name = parse_monthly_from_zip_bytes(raw)
                meta = {
                    "type": "inegi_xlsx_zip",
                    "source_url": url,
                    "zip_member": inner_name,
                }
            else:
                rows = parse_monthly_from_xlsx_bytes(raw)
                meta = {
                    "type": "inegi_xlsx",
                    "source_url": url,
                    "zip_member": "",
                }
        except Exception as exc:  # noqa: BLE001
            if verbose:
                print(f"WARN xlsx candidate failed {url}: {exc}")
            continue

        max_key = max((r.year, r.month) for r in rows)
        if max_key > best_key:
            best_rows = rows
            best_meta = meta
            best_key = max_key

    if best_rows is None or best_meta is None:
        raise RuntimeError("No valid INEGI xlsx candidates parsed")

    return best_rows, best_meta


def main() -> int:
    args = parse_args()
    repo_root = Path(args.repo_root).resolve()
    doc_dir = Path(args.doc_dir).resolve()
    state_dir = Path(args.state_dir).resolve()
    state_dir.mkdir(parents=True, exist_ok=True)

    if (args.target_year is None) ^ (args.target_quarter is None):
        print("ERROR: provide both --target-year and --target-quarter, or neither.")
        return 2

    run_id = timestamp_slug()
    summary_path = state_dir / f"phase1b_poverty_sync_{run_id}.json"
    monthly_csv = doc_dir / "poverty_lines_monthly.csv"
    quarterly_csv = doc_dir / "poverty_lines_quarterly.csv"

    summary: dict[str, Any] = {
        "version": 1,
        "run_id": run_id,
        "timestamp_utc": utc_now_iso(),
        "config": {
            "repo_root": str(repo_root),
            "doc_dir": str(doc_dir),
            "state_dir": str(state_dir),
            "source_mode": args.source_mode,
            "target_year": args.target_year,
            "target_quarter": args.target_quarter,
            "dry_run": args.dry_run,
            "skip_do_patch": args.skip_do_patch,
        },
        "paths": {
            "monthly_csv": str(monthly_csv),
            "quarterly_csv": str(quarterly_csv),
        },
        "status": "running",
    }

    try:
        monthly_rows: list[MonthlyPoint] = []
        source_meta: dict[str, Any] = {}
        xlsx_rows: list[MonthlyPoint] | None = None
        xlsx_meta: dict[str, Any] | None = None
        api_rows: list[MonthlyPoint] | None = None
        api_meta: dict[str, Any] | None = None

        if args.source_mode in {"auto", "xlsx"}:
            candidates = discover_inegi_xlsx_candidates(timeout_seconds=args.timeout_seconds, verbose=args.verbose)
            summary["xlsx_candidates"] = candidates
            if not candidates and args.source_mode == "xlsx":
                raise RuntimeError("No INEGI xlsx/zip candidates discovered on LP sources")
            if candidates:
                try:
                    xlsx_rows, xlsx_meta = choose_best_monthly_from_xlsx(
                        candidates,
                        timeout_seconds=args.timeout_seconds,
                        verbose=args.verbose,
                    )
                except Exception as exc:  # noqa: BLE001
                    summary["xlsx_error"] = str(exc)
                    if args.source_mode == "xlsx":
                        raise

        if args.source_mode in {"auto", "api"}:
            try:
                api_rows, api_meta = monthly_from_inegi_api(timeout_seconds=args.timeout_seconds)
            except Exception as exc:  # noqa: BLE001
                summary["api_error"] = str(exc)
                if args.source_mode == "api":
                    raise

        if args.source_mode == "xlsx":
            if not xlsx_rows or not xlsx_meta:
                raise RuntimeError("XLSX mode requested but no valid INEGI xlsx data was parsed")
            monthly_rows, source_meta = xlsx_rows, xlsx_meta
            summary["source_selection"] = {"mode": "xlsx", "reason": "user_forced_xlsx"}
        elif args.source_mode == "api":
            if not api_rows or not api_meta:
                raise RuntimeError("API mode requested but INEGI indicator API returned no valid series")
            monthly_rows, source_meta = api_rows, api_meta
            summary["source_selection"] = {"mode": "api", "reason": "user_forced_api"}
        else:
            if xlsx_rows and xlsx_meta and api_rows and api_meta:
                xlsx_key = max_month_key(xlsx_rows)
                api_key = max_month_key(api_rows)
                if api_key >= xlsx_key:
                    monthly_rows, source_meta = api_rows, api_meta
                    summary["source_selection"] = {
                        "mode": "api",
                        "reason": "newer_or_equal_coverage",
                        "xlsx_latest": {"year": xlsx_key[0], "month": xlsx_key[1]},
                        "api_latest": {"year": api_key[0], "month": api_key[1]},
                    }
                else:
                    monthly_rows, source_meta = xlsx_rows, xlsx_meta
                    summary["source_selection"] = {
                        "mode": "xlsx",
                        "reason": "newer_coverage",
                        "xlsx_latest": {"year": xlsx_key[0], "month": xlsx_key[1]},
                        "api_latest": {"year": api_key[0], "month": api_key[1]},
                    }
            elif xlsx_rows and xlsx_meta:
                monthly_rows, source_meta = xlsx_rows, xlsx_meta
                summary["source_selection"] = {"mode": "xlsx", "reason": "api_unavailable"}
            elif api_rows and api_meta:
                monthly_rows, source_meta = api_rows, api_meta
                summary["source_selection"] = {"mode": "api", "reason": "xlsx_unavailable"}
            else:
                monthly_rows = read_monthly_csv(monthly_csv)
                source_meta = {
                    "type": "local_cached_csv",
                    "source_file": str(monthly_csv),
                }
                summary["source_selection"] = {
                    "mode": "local_cached_csv",
                    "reason": "xlsx_and_api_unavailable",
                }

        quarterly_rows = build_quarterly(monthly_rows)
        complete_quarters = [q for q in quarterly_rows if q.month_count == 3]

        summary["source"] = source_meta
        summary["rows"] = {
            "monthly": len(monthly_rows),
            "quarterly_complete": len(complete_quarters),
            "quarterly_total": len(quarterly_rows),
            "latest_month": {
                "year": monthly_rows[-1].year,
                "month": monthly_rows[-1].month,
            },
        }

        if args.target_year is not None and args.target_quarter is not None:
            target_row = next(
                (
                    q
                    for q in complete_quarters
                    if q.year == args.target_year and q.quarter == args.target_quarter
                ),
                None,
            )
            if target_row is None:
                raise RuntimeError(
                    f"Target quarter {args.target_year}-Q{args.target_quarter} missing in complete quarterly series"
                )
            summary["target_values"] = {
                "year": target_row.year,
                "quarter": target_row.quarter,
                "rural": target_row.rural,
                "urban": target_row.urban,
            }

        if not args.dry_run:
            doc_dir.mkdir(parents=True, exist_ok=True)
            write_monthly_csv(monthly_csv, monthly_rows)
            write_quarterly_csv(quarterly_csv, complete_quarters)

        patch_result = {"status": "skipped"}
        if args.target_year is not None and args.target_quarter is not None and not args.skip_do_patch:
            patch_result = patch_target_harmonization_do(
                repo_root=repo_root,
                year=args.target_year,
                quarter=args.target_quarter,
                dry_run=args.dry_run,
            )
        summary["do_patch"] = patch_result

        summary["status"] = "ok"
        if args.dry_run:
            summary["status"] = "would_update"

    except Exception as exc:  # noqa: BLE001
        summary["status"] = "failed"
        summary["error"] = str(exc)

    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    print(f"Phase 1B poverty sync status: {summary['status']}")
    print(f"Summary: {summary_path}")
    if "target_values" in summary:
        tv = summary["target_values"]
        print(f"Target {tv['year']}-Q{tv['quarter']}: rural={tv['rural']:.2f}, urban={tv['urban']:.2f}")
    if "do_patch" in summary:
        print(f"do_patch: {summary['do_patch'].get('status', 'n/a')}")

    return 0 if summary["status"] in {"ok", "would_update"} else 1


if __name__ == "__main__":
    raise SystemExit(main())
